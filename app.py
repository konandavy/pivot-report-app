import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openai import OpenAI
import plotly.express as px

st.set_page_config(page_title="Pivot Report Generator", layout="wide")
st.title("📊 Automated Pivot Report Generator by Konan Davy")

# File uploader
uploaded_file = st.file_uploader("Upload your Excel file with aggregate data", type=[".xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file, sheet_name='Aggregate Data')

    # Convert Time to numeric and compute Hours
    df['Time'] = pd.to_numeric(df['Time'], errors='coerce')
    df['Hours'] = df['Time'] / 60

    # Extract team members
    team_members = df['Source.Name'].dropna().unique()

    st.success(f"File loaded successfully. Found {len(team_members)} team members.")

    # Preview data
    if st.checkbox("Preview raw data"):
        st.dataframe(df.head(20))

    # === AI-generated Insights ===
    if st.checkbox("🤖 Show AI-generated insights"):
        top_client = df.groupby('Client')['Hours'].sum().idxmax()
        top_activity = df.groupby('Activity Name')['Hours'].sum().idxmax()
        top_member = df.groupby('Source.Name')['Hours'].sum().idxmax()
        max_week = df.groupby('Week')['Hours'].sum().idxmax()
        zero_hours = df[df['Hours'] == 0]

        st.markdown("### 🔍 Key Insights")
        st.markdown(f"- 🥇 **Top client by hours:** {top_client}")
        st.markdown(f"- 🛠️ **Top activity by hours:** {top_activity}")
        st.markdown(f"- 👤 **Top team member by hours:** {top_member}")
        st.markdown(f"- 📅 **Week with most logged hours:** {max_week}")
        st.markdown(f"- ⚠️ **Entries with 0 hours logged:** {len(zero_hours)}")

    # === Visualizations ===
    if st.checkbox("📈 Show Visual Charts"):
        st.subheader("Distribution of Hours by Client")
        client_hours = df.groupby("Client")["Hours"].sum().reset_index()
        fig_client = px.pie(client_hours, names='Client', values='Hours',
                            title='Distribution of Hours by Client', hole=0.3)
        st.plotly_chart(fig_client, use_container_width=True)

        st.subheader("Distribution of Hours by Activity")
        activity_hours = df.groupby("Activity Name")["Hours"].sum().reset_index()
        fig_activity = px.pie(activity_hours, names='Activity Name', values='Hours',
                              title='Distribution of Hours by Activity', hole=0.3)
        st.plotly_chart(fig_activity, use_container_width=True)

    # === Natural Language Q&A ===
    if st.checkbox("💬 Ask questions about the data"):
        question = st.text_input("Ask me anything about this dataset:")
        if question:
            import openai
            context_csv = df.head(100).to_csv(index=False)
            client = OpenAI(api_key=st.secrets["openai_api_key"])

            prompt = f"You are a data expert. Here's a dataset:\n{context_csv}\n\nQuestion: {question}\nAnswer:"

            try:
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are a helpful data expert."},
                        {"role": "user", "content": prompt}
                    ]
                )
                answer = response.choices[0].message.content
                st.markdown("### 🤖 Answer")
                st.write(answer)
            except Exception as e:
                st.error(f"Error generating answer: {e}")

    if st.button("Generate Pivot Report"):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill

            wb = writer.book
            ws = wb.create_sheet(title='Summary', index=0)

            def write_pivot_block(title, df_block):
                ws.append([title])
                ws.append(list(df_block.columns))
                for r in df_block.itertuples(index=False):
                    ws.append(list(r))
                ws.append([None])

            pivot1 = df.groupby('Client')['Hours'].sum().reset_index()
            pivot1['Hours'] = pivot1['Hours'].round(2)
            pivot1.columns = ['Row Labels', 'Sum of Hours']
            write_pivot_block('Client Totals', pivot1)

            pivot2 = df.groupby('Activity Name')['Hours'].sum().reset_index()
            pivot2['Hours'] = pivot2['Hours'].round(2)
            pivot2.columns = ['Row Labels', 'Sum of Hours']
            write_pivot_block('Activity Totals', pivot2)

            pivot3 = df.groupby('Week')['Hours'].sum().reset_index()
            pivot3['Hours'] = pivot3['Hours'].round(2)
            pivot3.columns = ['Row Labels', 'Sum of Hours']
            write_pivot_block('Week Totals', pivot3)

            grand_total = round(df['Hours'].sum(), 2)
            ws.append(['Grand Total', grand_total])

            bold_font = Font(bold=True)
            fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            for row in ws.iter_rows():
                if row[0].value in ['Client Totals', 'Activity Totals', 'Week Totals', 'Grand Total']:
                    for cell in row:
                        if cell.value is not None:
                            cell.font = bold_font
                            cell.fill = fill

            for member in team_members:
                member_df = df[df['Source.Name'] == member]
                detailed_log = member_df[['Client', 'Week', 'Activity Name', 'Comments', 'Time', 'Hours']].sort_values(by=['Client', 'Week'])
                detailed_log.to_excel(writer, sheet_name=f"{member[:15]}_Details", index=False)

        output.seek(0)
        st.download_button(
            label="📥 Download Pivot Report",
            data=output,
            file_name=f"pivot_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


